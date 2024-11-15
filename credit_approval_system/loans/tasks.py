from celery import shared_task
import openpyxl
import os

@shared_task
def process_loan_data(file_path):
    """Process loan data from uploaded Excel file."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Processing loan: {row}")

@shared_task
def process_customer_data(file_path):
    """Process customer data from uploaded Excel file."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Processing customer: {row}")
